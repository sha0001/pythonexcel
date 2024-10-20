# Manipulating the content of a workbook

import openpyxl
from openpyxl.comments import Comment
from collections import defaultdict

#Create a new workbook
filename = "FinancialSample.xlsx"

#Load the workbook
wb = openpyxl.load_workbook(filename)

#Get the active worksheet
sheet = wb.active

# Get entire column or row of cells

col = sheet["C"]
row = sheet[10]
print(f"{len(col)} cells in column")
print(f"{len(row)} cells in row")

# Get a range of cells

range = sheet["A2:B7"]
print(f"{len(range)} cells in range")
print(range)

# Iterate over rows and columns

for col in sheet.iter_cols(min_row=2, max_row=3, min_col=2, max_col=5):
    for cell in col:
        print(cell.value)

# Counting how many times each value appears in a row

counter = defaultdict(int)
for row in sheet.iter_rows(min_row=2, min_col=3, max_col=3):
    for cell in row:
        counter[cell.value] += 1
print(counter)

# Create a cell with a comment in it
# Note that this comment format is a bit buggy in online excel
# Seems to work fine in regular excel
cell = sheet["A1"]
cell.comment = Comment("This is a comment", "Liam Shannon")

# Save the workbook
wb.save("Content.xlsx")
