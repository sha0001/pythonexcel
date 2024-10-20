
# LinkedIn Learning Course
# Example file for Python: Working with Excel and Spreadsheet Data by Joe Marini
# Split a single worksheet into multiple worksheets

import openpyxl
from openpyxl.utils.cell import column_index_from_string


def split_workbook(workbook, source_sheet_name, split_column):
    source_worksheet = workbook[source_sheet_name]
# Creating an empty set to track sheet names.
# Sets can only contain unique values, so no chance of duplicates
    new_sheets = set()

# Creating a variable to track which sheet we are currently in 
    current_worksheet = None

    for row in source_worksheet.iter_rows(min_row=2):  # Skip header row

        # Get value from the specified column. this function converts A,B,C... column headings to a numerical index 
        col_indx = column_index_from_string(split_column) - 1
        value = row[col_indx].value

        # Adding the country value to the new sheets if its not already there. 
        if value not in new_sheets:
            new_sheets.add(value)

            # Create a new worksheet for the new value
            current_worksheet = workbook.create_sheet(title=value)
        else:
            current_worksheet = workbook[value]

        # Copy the row to the appropriate worksheet
        newrow = []
        for cell in row:
            newrow.append(cell.value)

        current_worksheet.append(newrow)

    # When complete, return the list of sheets that were added
    return new_sheets


# Example usage
filename = "FinancialSample.xlsx"
wb = openpyxl.load_workbook(filename)

# The split_workbook function allows for switching "B" to any other column.

source_sheet_name = "SalesData"
added_sheets = split_workbook(wb, source_sheet_name, "B")

# Add the auto-filters to each worksheet
for sheet_name in added_sheets:
    sheet = wb[sheet_name]
    filters = sheet.auto_filter
    filters.ref = sheet.dimensions

# when the loop completes, save the new sheet
wb.save("new"+filename)
