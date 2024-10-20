# Applying conditional formatting to a worksheet

import openpyxl
from openpyxl.formatting import Rule
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.differential import DifferentialStyle

filename = "FinancialSample.xlsx"

# Load the workbook
workbook = openpyxl.load_workbook(filename)
sheet = workbook["SalesData"]

# Define the style to represent the formatting
red_colour = "ffd2d2"
bold_text = Font(bold=True, color ="00ff0000")
red_fill = PatternFill(bgColor= red_colour, fill_type = "solid")

diff_style = DifferentialStyle(font=bold_text, fill=red_fill)

# Create a rule for the conditional formatting
rule = Rule(type="expression",dxf=diff_style, formula=["$L1<10000"])

# Add the rule to the entire sheet
dimensions = "L1:L701"
sheet.conditional_formatting.add(dimensions, rule)

workbook.save("CondFormat.xlsx")
print("Workbook created successfuly!") 

