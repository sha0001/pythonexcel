# Exploring an xlsx file with openpyxl

import openpyxl 

filename = "FinancialSample.xlsx" 

#Load the workbook
workbook = openpyxl.load_workbook(filename)

#Print Basic information
print(f"Number of worksheets: {len(workbook.sheetnames)}")
for worksheet_name in workbook.sheetnames:
    worksheet = workbook[worksheet_name]
    print(f"\nWorksheet: {worksheet_name}")
    #Explore each worksheet

    # Get Dimensions 
    dimensions = worksheet.dimensions
    print(f"    - Dimensions: {dimensions}")

    print(f"Min row: {worksheet.min_row}")
    print(f"Max row: {worksheet.max_row}")

    print(f"Min column: {worksheet.min_column}")
    print(f"Max column: {worksheet.max_column}")

    #Check if worksheet is empty
    if worksheet.max_row == 1 and worksheet.max_column == 1:
       print('Worksheet is empty')
    else: 
       cell = worksheet["A1"]
       print(f"   - Top-left cell value: {cell.value}")
       cell = worksheet.cell(row=worksheet.max_row, column = worksheet.max_column)
       print(f"   - Bottom-right cell value: {cell.value}")


